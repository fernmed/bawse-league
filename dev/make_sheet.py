"""Bawse League Championship League draft workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from teams_data import LEAGUES, ODDS

ACCENT = "1F4E79"
LIGHT = "EDF2F8"
YELLOW = "FFF2CC"
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HDR_FILL = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
ALT_FILL = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
INPUT_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=16, bold=True, color=ACCENT)
BASE = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
ITAL = Font(name="Arial", size=9, italic=True, color="595959")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

N_MGRS, N_RDS = 10, 11
BOARD = "'Draft Board'!$B$3:$K$13"

wb = openpyxl.Workbook()

# ---------------- Master Team List ----------------
ws_m = wb.active
ws_m.title = "Master Team List"
for c, (h, w) in enumerate([("League", 10), ("Team", 28), ("Draft Name (dropdown source)", 34),
                            ("Drafted?", 10)], start=1):
    cell = ws_m.cell(row=1, column=c, value=h)
    cell.font, cell.fill = HDR_FONT, HDR_FILL
    ws_m.column_dimensions[get_column_letter(c)].width = w

rows = []
for lg, teams, champ in LEAGUES:
    for t in teams:
        rows.append((lg, t, f"{lg}: {t}"))

for i, (lg, team, dn) in enumerate(rows, start=2):
    ws_m.cell(row=i, column=1, value=lg).font = BASE
    ws_m.cell(row=i, column=2, value=team).font = BASE
    ws_m.cell(row=i, column=3, value=dn).font = BASE
    d = ws_m.cell(row=i, column=4, value=f'=IF(COUNTIF({BOARD},C{i})>0,"DRAFTED","")')
    d.font = BOLD
last_row = 1 + len(rows)
ws_m.freeze_panes = "A2"

# ---------------- Draft Board ----------------
ws = wb.create_sheet("Draft Board")
ws.sheet_view.showGridLines = False
t = ws.cell(row=1, column=1, value="BAWSE LEAGUE — CHAMPIONSHIP LEAGUE DRAFT BOARD")
t.font = TITLE_FONT
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

ws.cell(row=2, column=1, value="Round").font = HDR_FONT
ws.cell(row=2, column=1).fill = HDR_FILL
for m in range(N_MGRS):
    c = ws.cell(row=2, column=2 + m, value=f"Manager {m + 1}")
    c.font, c.fill, c.border = BOLD, INPUT_FILL, THIN
    c.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(2 + m)].width = 26
ws.column_dimensions["A"].width = 9

for r in range(N_RDS):
    arrow = "→" if r % 2 == 0 else "←"  # snake direction
    lab = ws.cell(row=3 + r, column=1, value=f"{r + 1} {arrow}")
    lab.font, lab.fill = HDR_FONT, HDR_FILL
    lab.alignment = Alignment(horizontal="center")
    for m in range(N_MGRS):
        cell = ws.cell(row=3 + r, column=2 + m)
        cell.font, cell.border = BASE, THIN
        if r % 2 == 1:
            cell.fill = ALT_FILL

# picks-made row
ws.cell(row=14, column=1, value="Picks").font = ITAL
for m in range(N_MGRS):
    col = get_column_letter(2 + m)
    c = ws.cell(row=14, column=2 + m, value=f"=COUNTA({col}3:{col}13)")
    c.font, c.alignment = ITAL, Alignment(horizontal="center")

ws.cell(row=15, column=2,
        value="Snake draft: odd rounds left→right, even rounds right←left. "
              "A pick that turns RED is a duplicate — re-pick immediately.").font = ITAL

# dropdown validation
dv = DataValidation(type="list", formula1=f"='Master Team List'!$C$2:$C${last_row}",
                    allow_blank=True, showErrorMessage=True,
                    error="Pick a team from the Master Team List dropdown.")
ws.add_data_validation(dv)
dv.add("B3:K13")

# duplicate highlighting
ws.conditional_formatting.add(
    "B3:K13",
    FormulaRule(formula=['AND(B3<>"",COUNTIF($B$3:$K$13,B3)>1)'], fill=RED_FILL, stopIfTrue=False),
)

# league tracker
tr = 17
c = ws.cell(row=tr, column=1, value="LIVE DRAFT TRACKER")
c.font = Font(name="Arial", size=12, bold=True, color=ACCENT)
hdrs = ["League", "Teams", "Drafted", "Undrafted"]
for j, h in enumerate(hdrs):
    cell = ws.cell(row=tr + 1, column=1 + j, value=h)
    cell.font, cell.fill, cell.border = HDR_FONT, HDR_FILL, THIN
for i, (lg, teams, champ) in enumerate(LEAGUES):
    r = tr + 2 + i
    vals = [
        lg,
        len(teams),
        f'=COUNTIF({BOARD},"{lg}: *")',
        f'=B{r}-C{r}',
    ]
    for j, v in enumerate(vals):
        cell = ws.cell(row=r, column=1 + j, value=v)
        cell.font, cell.border = BASE, THIN
        if i % 2 == 1:
            cell.fill = ALT_FILL
tot = tr + 2 + len(LEAGUES)
ws.cell(row=tot, column=1, value="TOTAL").font = BOLD
ws.cell(row=tot, column=2, value=f"=SUM(B{tr+2}:B{tot-1})").font = BOLD
ws.cell(row=tot, column=3, value=f"=SUM(C{tr+2}:C{tot-1})").font = BOLD
ws.cell(row=tot, column=4, value=f"=SUM(D{tr+2}:D{tot-1})").font = BOLD
ws.cell(row=tot + 1, column=1,
        value='"Drafted" counts drafted teams; undrafted teams stay unowned for the season.').font = ITAL
ws.freeze_panes = "B3"

# ---------------- Odds Cheatsheet ----------------
ws_o = wb.create_sheet("Odds Cheatsheet")
ws_o.sheet_view.showGridLines = False
t = ws_o.cell(row=1, column=1, value="CHAMPIONSHIP FUTURES CHEATSHEET")
t.font = TITLE_FONT
ws_o.cell(row=2, column=1,
          value="Odds as of July 2026 (MLB All-Star break) — current for a late-July 2026 "
                "draft. Spot-check lines the week of draft night.").font = ITAL
for col, w in [("A", 30), ("B", 10), ("C", 12), ("E", 30), ("F", 10), ("G", 12)]:
    ws_o.column_dimensions[col].width = w

def odds_block(ws_o, top, left, lg, title, book, entries):
    c = ws_o.cell(row=top, column=left, value=f"{lg} — {title}")
    c.font = Font(name="Arial", size=11, bold=True, color=ACCENT)
    ws_o.cell(row=top + 1, column=left, value=f"Book: {book}").font = ITAL
    for j, h in enumerate(["Team", "Odds", "Implied %"]):
        cell = ws_o.cell(row=top + 2, column=left + j, value=h)
        cell.font, cell.fill, cell.border = HDR_FONT, HDR_FILL, THIN
    for i, (team, odds) in enumerate(entries):
        r = top + 3 + i
        a = ws_o.cell(row=r, column=left, value=team)
        b = ws_o.cell(row=r, column=left + 1, value=odds)
        b.number_format = '"+"#,##0'
        col_b = get_column_letter(left + 1)
        p = ws_o.cell(row=r, column=left + 2, value=f"=100/({col_b}{r}+100)")
        p.number_format = "0.0%"
        for cell in (a, b, p):
            cell.font, cell.border = BASE, THIN
            if i % 2 == 1:
                cell.fill = ALT_FILL
    return top + 3 + len(entries)

order = ["MLB", "NBA", "NHL", "NFL", "NCAAF", "NCAAM", "EPL"]
tops = {0: 4, 1: 4}
for idx, lg in enumerate(order):
    side = idx % 2
    title, book, entries = ODDS[lg]
    tops[side] = odds_block(ws_o, tops[side], 1 if side == 0 else 5, lg, title, book, entries) + 2

# ---------------- How To Use ----------------
ws_h = wb.create_sheet("How To Use")
ws_h.sheet_view.showGridLines = False
ws_h.column_dimensions["A"].width = 110
t = ws_h.cell(row=1, column=1, value="HOW TO USE THIS DRAFT TOOL")
t.font = TITLE_FONT
lines = [
    ("", ""),
    ("1. Before the draft", "bold"),
    ("• Type your 10 manager names into the yellow cells on the Draft Board (row 2).", ""),
    ("• Randomize draft order; the column order IS the draft order.", ""),
    ("• The Odds Cheatsheet is current as of July 2026 — spot-check big movers the week of the draft.", ""),
    ("• Update the Master Team List for renames, relocation, EPL promotion/relegation, and FBS/D1 changes.", ""),
    ("", ""),
    ("2. During the draft", "bold"),
    ("• On each pick, use the dropdown in your cell for that round. Odd rounds go left→right, even rounds right←left (snake).", ""),
    ("• Teams are listed as “LEAGUE: Team”. Type the league prefix to jump there in the dropdown.", ""),
    ("• A duplicate pick turns RED in both cells — the later pick is void and must be re-made.", ""),
    ("• The Live Draft Tracker (below the board) shows how many teams remain undrafted in each league.", ""),
    ("• The Master Team List marks every individually drafted team with “DRAFTED”.", ""),
    ("", ""),
    ("3. After the draft", "bold"),
    ("• Each manager’s roster is simply their column, top to bottom.", ""),
    ("• Log trades by editing the cells and noting the swap in your group chat.", ""),
    ("• Save a copy each season — next year’s update takes about 75 minutes (per the format’s inventor).", ""),
    ("", ""),
    ("Format credit: adapted from a Championship League described on r/FFCommish (u/Beaushekai).", "ital"),
    ("Cells you should edit: yellow. Everything with formulas updates automatically.", "ital"),
]
for i, (txt, style) in enumerate(lines, start=2):
    c = ws_h.cell(row=i, column=1, value=txt)
    c.font = BOLD if style == "bold" else (ITAL if style == "ital" else BASE)
    c.alignment = Alignment(wrap_text=True, vertical="top")

wb._sheets = [wb["Draft Board"], wb["Master Team List"], wb["Odds Cheatsheet"], wb["How To Use"]]
wb.save("Bawse League - Draft Board.xlsx")
print("saved; master rows:", last_row - 1)
